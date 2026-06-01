import os
import json
import urllib.request
import urllib.parse
import re
import time
import traceback
from datetime import datetime
from flask import Flask, render_template, request, jsonify, send_file
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import tempfile
import requests as req_lib

# Load .env file (for Tencent Cloud deployment)
_env_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), '.env')
if os.path.exists(_env_file):
    with open(_env_file) as _f:
        for _line in _f:
            _line = _line.strip()
            if _line and not _line.startswith('#') and '=' in _line:
                _k, _, _v = _line.partition('=')
                os.environ.setdefault(_k.strip(), _v.strip())

# Template folder: use absolute path so it works on Vercel/serverless
_TEMPLATE_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'templates')
_STATIC_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'static')
app = Flask(__name__, template_folder=_TEMPLATE_DIR, static_folder=_STATIC_DIR)

# ==================== Config ====================
PROXY_PORT = os.environ.get('YT_PROXY_PORT', '7897')
PROXY_URL = f'http://127.0.0.1:{PROXY_PORT}'
USE_PROXY = os.environ.get('USE_PROXY', 'true').lower() == 'true'
APP_PORT = int(os.environ.get('APP_PORT', os.environ.get('PORT', '5146')))

# TikHub API
TIKHUB_API_KEY = os.environ.get('TIKHUB_API_KEY', '')
TIKHUB_API_BASE = 'https://api.tikhub.dev'

# LLM Config
LLM_API_KEY = os.environ.get('LLM_API_KEY', '')
LLM_API_URL = os.environ.get('LLM_API_URL', '')
LLM_MODEL = os.environ.get('LLM_MODEL', '')

# Settings file for persistence
_SETTINGS_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), '.settings.json')

def _load_settings():
    """Load settings from file on startup."""
    global TIKHUB_API_KEY, LLM_API_KEY, LLM_API_URL, LLM_MODEL
    try:
        if os.path.exists(_SETTINGS_FILE):
            with open(_SETTINGS_FILE, 'r') as f:
                s = json.load(f)
            if s.get('tikhub_key') and not TIKHUB_API_KEY:
                TIKHUB_API_KEY = s['tikhub_key']
            if s.get('llm_key') and not LLM_API_KEY:
                LLM_API_KEY = s['llm_key']
            if s.get('llm_api_url') and not LLM_API_URL:
                LLM_API_URL = s['llm_api_url']
            if s.get('llm_model') and not LLM_MODEL:
                LLM_MODEL = s['llm_model']
            print(f'✅ 设置已从文件加载 (TikHub: {"已配置" if TIKHUB_API_KEY else "未配置"}, LLM: {"已配置" if LLM_API_KEY else "未配置"})')
    except Exception as e:
        print(f'⚠️ 加载设置文件失败: {e}')

def _save_settings_to_file():
    """Save current settings to file."""
    try:
        with open(_SETTINGS_FILE, 'w') as f:
            json.dump({
                'tikhub_key': TIKHUB_API_KEY,
                'llm_key': LLM_API_KEY,
                'llm_api_url': LLM_API_URL,
                'llm_model': LLM_MODEL,
            }, f)
    except Exception as e:
        print(f'⚠️ 保存设置文件失败: {e}')

# Load settings on startup
_load_settings()

# CORS headers
@app.after_request
def after_request(response):
    response.headers.add('Access-Control-Allow-Origin', '*')
    response.headers.add('Access-Control-Allow-Headers', 'Content-Type')
    response.headers.add('Access-Control-Allow-Methods', 'GET,POST,OPTIONS')
    return response


# ==================== TikHub Error ====================
class TikHubError(Exception):
    pass


# ==================== TikHub Helpers ====================
def tikhub_get(path, api_key, params=None, timeout=30):
    url = f'{TIKHUB_API_BASE}{path}'
    if params:
        qs = urllib.parse.urlencode({k: v for k, v in params.items() if v is not None})
        if qs:
            url += f'?{qs}'
    r = req_lib.get(url, headers={
        'Authorization': f'Bearer {api_key}',
        'Content-Type': 'application/json',
    }, timeout=timeout, proxies={'http': PROXY_URL, 'https': PROXY_URL} if USE_PROXY else None)
    if r.status_code == 402:
        raise TikHubError('TikHub API 余额不足，请充值后重试')
    if r.status_code == 401:
        raise TikHubError('TikHub API Key 无效或已过期')
    if r.status_code != 200:
        # Try to extract error message from TikHub response
        try:
            err_data = r.json()
            detail = err_data.get('detail', err_data)
            if isinstance(detail, dict):
                err_msg = detail.get('message', '') or detail.get('message_zh', '')
            else:
                err_msg = str(detail)[:200]
        except:
            err_msg = ''
        raise TikHubError(f'TikHub API 错误: HTTP {r.status_code}' + (f' - {err_msg}' if err_msg else ''))
    return r.json()


def tikhub_post(path, api_key, body=None, timeout=30):
    url = f'{TIKHUB_API_BASE}{path}'
    r = req_lib.post(url, headers={
        'Authorization': f'Bearer {api_key}',
        'Content-Type': 'application/json',
    }, json=body or {}, timeout=timeout,
    proxies={'http': PROXY_URL, 'https': PROXY_URL} if USE_PROXY else None)
    if r.status_code == 402:
        raise TikHubError('TikHub API 余额不足，请充值后重试')
    if r.status_code == 401:
        raise TikHubError('TikHub API Key 无效或已过期')
    if r.status_code != 200:
        # Try to extract error message from TikHub response
        try:
            err_data = r.json()
            detail = err_data.get('detail', err_data)
            if isinstance(detail, dict):
                err_msg = detail.get('message', '') or detail.get('message_zh', '')
            else:
                err_msg = str(detail)[:200]
        except:
            err_msg = ''
        raise TikHubError(f'TikHub API 错误: HTTP {r.status_code}' + (f' - {err_msg}' if err_msg else ''))
    return r.json()


def format_number(n):
    """Format number to readable string."""
    try:
        n = int(n)
        if n >= 10000_0000:
            return f'{n/10000_0000:.1f}亿'
        if n >= 10000:
            return f'{n/10000:.1f}万'
        return str(n)
    except (ValueError, TypeError):
        return str(n)


# ==================== Platform: Douyin ====================
def douyin_search_user(keyword):
    """Search douyin user by keyword using video search and extracting unique authors.
    (fetch_user_search_v2 is unstable, so we use video search as fallback)"""
    result = tikhub_post('/api/v1/douyin/search/fetch_video_search_v1', TIKHUB_API_KEY,
                         body={'keyword': keyword, 'page': 1})
    items = result.get('data', {}).get('data', [])
    out = []
    seen_sec = set()
    for item in items:
        if not isinstance(item, dict):
            continue
        aweme = item.get('aweme_info', {})
        if not isinstance(aweme, dict):
            continue
        author = aweme.get('author', {})
        if not isinstance(author, dict):
            continue
        sec_uid = author.get('sec_uid', '')
        if not sec_uid or sec_uid in seen_sec:
            continue
        seen_sec.add(sec_uid)

        # Get avatar (prefer larger size)
        avatar = ''
        avatar_larger = author.get('avatar_larger', {})
        if isinstance(avatar_larger, dict):
            urls = avatar_larger.get('url_list', [])
            if urls:
                avatar = urls[0]
        if not avatar:
            avatar_thumb = author.get('avatar_thumb', {})
            if isinstance(avatar_thumb, dict):
                urls = avatar_thumb.get('url_list', [])
                if urls:
                    avatar = urls[0]
        if avatar and avatar.startswith('//'):
            avatar = 'https:' + avatar

        # unique_id is the Douyin handle (抖音号)
        unique_id = author.get('unique_id', '') or author.get('short_id', '')
        display_id = unique_id if unique_id else str(sec_uid)[:12] + '...'

        out.append({
            'id': sec_uid,  # sec_uid for API calls
            'name': author.get('nickname', ''),
            'avatar': avatar,
            'fans': author.get('follower_count', 0),
            'likes': author.get('total_favorited', 0),
            'posts': author.get('aweme_count', 0),
            'platform': 'douyin',
            'platform_id': display_id,
            'platform_id_label': '抖音号',
        })
    return out


def douyin_get_profile(sec_user_id):
    """Get douyin user profile."""
    result = tikhub_get('/api/v1/douyin/app/v3/handler_user_profile', TIKHUB_API_KEY,
                        params={'sec_user_id': sec_user_id})
    user = result.get('data', {}).get('user', {})
    avatar = user.get('avatar_larger', {}).get('url_list', [''])[0] if isinstance(user.get('avatar_larger'), dict) else ''
    if avatar and avatar.startswith('//'):
        avatar = 'https:' + avatar
    return {
        'id': sec_user_id,
        'name': user.get('nickname', ''),
        'avatar': avatar,
        'signature': user.get('signature', ''),
        'fans': user.get('follower_count', 0),
        'following': user.get('following_count', 0),
        'likes': user.get('total_favorited', 0),
        'posts': user.get('aweme_count', 0),
        'platform': 'douyin',
    }


def douyin_get_videos(sec_user_id, count=20):
    """Get douyin user's posted videos."""
    result = tikhub_get('/api/v1/douyin/app/v3/fetch_user_post_videos', TIKHUB_API_KEY,
                        params={'sec_user_id': sec_user_id, 'count': count, 'max_cursor': 0})
    aweme_list = result.get('data', {}).get('aweme_list', []) or []
    videos = []
    for v in aweme_list:
        stats = v.get('statistics', {})
        music = v.get('music', {}) or {}
        desc = v.get('desc', '')
        # Clean hashtags from desc
        text_extra = v.get('text_extra') or []
        hashtags = [t.get('hashtag_name', '') for t in text_extra if isinstance(t, dict) and t.get('hashtag_name')]

        # Duration in seconds
        duration_ms = v.get('duration', 0)
        try:
            duration_sec = int(duration_ms) // 1000 if duration_ms else 0
        except (ValueError, TypeError):
            duration_sec = 0

        cover_url = ''
        cover = v.get('video', {}).get('cover')
        if isinstance(cover, dict):
            url_list = cover.get('url_list', [])
            cover_url = url_list[0] if url_list else ''
        elif isinstance(cover, str):
            cover_url = cover
        # Fix protocol-relative URLs
        if cover_url and cover_url.startswith('//'):
            cover_url = 'https:' + cover_url

        videos.append({
            'id': v.get('aweme_id', ''),
            'title': desc[:300],
            'thumbnail': cover_url,
            'viewCount': stats.get('play_count', 0),
            'likeCount': stats.get('digg_count', 0),
            'commentCount': stats.get('comment_count', 0),
            'shareCount': stats.get('share_count', 0),
            'collectCount': stats.get('collect_count', 0),
            'duration': duration_sec,
            'durationFormatted': f'{duration_sec // 60}:{duration_sec % 60:02d}',
            'publishedAt': datetime.fromtimestamp(v.get('create_time', 0)).strftime('%Y-%m-%d %H:%M') if v.get('create_time') else '',
            'music_title': music.get('title', ''),
            'music_author': music.get('author', ''),
            'music_id': music.get('id', ''),
            'hashtags': hashtags[:10],
            'is_top': v.get('is_top', 0) == 1,
            'url': f'https://www.douyin.com/video/{v.get("aweme_id", "")}',
            'platform': 'douyin',
        })
    return videos


# ==================== Platform: Xiaohongshu ====================
def xiaohongshu_search_user(keyword):
    """Search xiaohongshu user by keyword."""
    result = tikhub_get('/api/v1/xiaohongshu/app_v2/search_users', TIKHUB_API_KEY,
                        params={'keyword': keyword, 'page': 1})
    users = result.get('data', {}).get('data', {}).get('users', [])
    out = []
    for u in users[:10]:
        red_id = u.get('red_id', '')
        out.append({
            'id': u.get('id', ''),
            'name': u.get('name', ''),
            'avatar': u.get('image', ''),
            'fans': u.get('sub_title', ''),  # e.g. "Fans 334.7k"
            'desc': u.get('desc', ''),
            'red_id': red_id,
            'platform': 'xiaohongshu',
            'platform_id': red_id or str(u.get('id', '')),
            'platform_id_label': '小红书号',
        })
    return out


def xiaohongshu_get_profile(user_id):
    """Get xiaohongshu user profile."""
    result = tikhub_get('/api/v1/xiaohongshu/app_v2/get_user_info', TIKHUB_API_KEY,
                        params={'user_id': user_id})
    d = result.get('data', {})
    user = d.get('data', d)
    # Navigate nested structure (app_v2 returns flat user data, not nested under 'user')
    if isinstance(user, dict) and 'user' in user:
        user = user['user']
    avatar = user.get('imageb', '') or user.get('image', '') or user.get('avatar', '')
    return {
        'id': user_id,
        'name': user.get('nickname', user.get('name', '')),
        'avatar': avatar,
        'signature': user.get('desc', user.get('signature', '')),
        'fans': user.get('fans', user.get('fansCount', 0)),
        'following': user.get('follows', 0),
        'likes': user.get('liked', user.get('interaction', {}).get('likeCount', 0) if isinstance(user.get('interaction'), dict) else 0),
        'collected': user.get('collected', 0),
        'posts': user.get('notes', user.get('noteCount', 0)),
        'red_id': user.get('red_id', ''),
        'ip_location': user.get('ip_location', ''),
        'platform': 'xiaohongshu',
    }


def xiaohongshu_get_notes(user_id, count=20):
    """Get xiaohongshu user's posted notes."""
    result = tikhub_get('/api/v1/xiaohongshu/app_v2/get_user_posted_notes', TIKHUB_API_KEY,
                        params={'user_id': user_id, 'cursor': ''})
    items = result.get('data', {}).get('data', {}).get('notes', [])
    if not items:
        items = result.get('data', {}).get('data', {}).get('items', [])
    notes = []
    for n in items[:count]:
        # API v2 returns flat structure (no note_card / interact_info nesting)
        note = n if isinstance(n, dict) else {}
        if not note:
            continue

        # Get cover/thumbnail from images_list
        thumbnail = ''
        images_list = note.get('images_list', [])
        if images_list and isinstance(images_list, list) and isinstance(images_list[0], dict):
            thumbnail = images_list[0].get('url', '') or images_list[0].get('url_size_large', '')

        # Get interaction counts from flat fields
        note_id = note.get('note_id', note.get('id', ''))
        notes.append({
            'id': note_id,
            'title': note.get('title', note.get('display_title', '')),
            'thumbnail': thumbnail,
            'viewCount': note.get('view_count', 0),
            'likeCount': note.get('likes', 0),
            'commentCount': note.get('comments_count', 0),
            'shareCount': note.get('share_count', 0),
            'collectCount': note.get('collected_count', 0),
            'duration': 0,
            'durationFormatted': '',
            'publishedAt': note.get('time', note.get('last_update_time', '')),
            'music_title': '',
            'music_author': '',
            'music_id': '',
            'hashtags': [t.get('name', '') for t in (note.get('tag_list') or []) if isinstance(t, dict)][:10],
            'is_top': False,
            'url': f'https://www.xiaohongshu.com/explore/{note_id}',
            'type': note.get('type', ''),  # video or normal
            'platform': 'xiaohongshu',
        })
    return notes


# ==================== Platform: Bilibili ====================
def bilibili_search_user(keyword):
    """Search bilibili user - use general search and extract unique authors from video results,
    then fetch real profile data (including fans) for each author."""
    result = tikhub_get('/api/v1/bilibili/web/fetch_general_search', TIKHUB_API_KEY,
                        params={'keyword': keyword, 'order': '0', 'page': 1, 'page_size': 10})
    items = result.get('data', {}).get('data', {}).get('result', [])
    out = []
    seen_mids = set()
    for item in items:
        if not isinstance(item, dict):
            continue
        # Each item is a video result with author info
        mid = item.get('mid', '')
        if mid and str(mid) not in seen_mids:
            seen_mids.add(str(mid))
            author = item.get('author', '')
            upic = item.get('upic', '')
            if upic and upic.startswith('//'):
                upic = 'https:' + upic
            out.append({
                'id': str(mid),
                'name': author,
                'avatar': upic,
                'fans': 0,  # Will be updated below
                'platform': 'bilibili',
                'platform_id': str(mid),
                'platform_id_label': 'UID',
            })

    # Fetch real profile data for each author (to get actual fans count)
    for user in out[:5]:  # Limit to top 5 to avoid too many API calls
        try:
            uid = user['id']
            profile = bilibili_get_profile(uid)
            user['fans'] = profile.get('fans', 0)
            # Also update avatar with higher quality from profile
            if profile.get('avatar'):
                user['avatar'] = profile['avatar']
        except Exception as e:
            print(f'⚠️ B站获取用户 {user.get("name")} 粉丝数失败: {e}')
            # Keep fans as 0, don't block the whole search

    return out


def bilibili_get_profile(uid):
    """Get bilibili user profile."""
    # Get basic profile info
    try:
        result = tikhub_get('/api/v1/bilibili/web/fetch_user_profile', TIKHUB_API_KEY,
                            params={'uid': uid})
        d = result.get('data', {}).get('data', {})
        # New API structure: data is directly in data.data (no 'card' wrapper)
        name = d.get('name', '')
        avatar = d.get('face', '')
        signature = d.get('sign', '')
        level = d.get('level', 0)
        official = d.get('official', {})
        official_desc = official.get('title', '') if isinstance(official, dict) else ''
    except Exception as e:
        print(f'⚠️ B站获取用户 {uid} 基础资料失败: {e}')
        name = ''
        avatar = ''
        signature = ''
        level = 0
        official_desc = ''

    # Get follower count from relation_stat API (profile API no longer returns fans)
    fans = 0
    following = 0
    try:
        stat = tikhub_get('/api/v1/bilibili/web/fetch_user_relation_stat', TIKHUB_API_KEY,
                          params={'uid': uid})
        stat_data = stat.get('data', {})
        fans = stat_data.get('follower', 0)
        following = stat_data.get('following', 0)
    except Exception as e:
        print(f'⚠️ B站获取用户 {uid} 粉丝数失败: {e}')

    return {
        'id': str(uid),
        'name': name,
        'avatar': avatar,
        'signature': signature,
        'fans': fans,
        'following': following,
        'likes': 0,  # Not available from current API
        'posts': 0,  # Not available from current API
        'level': level,
        'official_verify': official_desc,
        'platform': 'bilibili',
    }


def bilibili_get_videos(uid, count=20):
    """Get bilibili user's posted videos with full stats.
    fetch_user_post_videos doesn't return likes, so we fetch video detail for each."""
    result = tikhub_get('/api/v1/bilibili/web/fetch_user_post_videos', TIKHUB_API_KEY,
                        params={'uid': uid, 'pn': 1, 'order': 'pubdate'})
    vlist = result.get('data', {}).get('data', {}).get('list', {}).get('vlist', [])
    if not vlist:
        vlist = result.get('data', {}).get('data', {}).get('vlist', [])

    videos = []
    for v in vlist[:count]:
        if not isinstance(v, dict):
            continue

        bvid = v.get('bvid', '')
        aid = v.get('aid', '')
        title = v.get('title', '')

        # Parse duration
        duration_str = v.get('length', '')
        duration_sec = 0
        if duration_str and ':' in str(duration_str):
            parts = str(duration_str).split(':')
            try:
                if len(parts) == 2:
                    duration_sec = int(parts[0]) * 60 + int(parts[1])
                elif len(parts) == 3:
                    duration_sec = int(parts[0]) * 3600 + int(parts[1]) * 60 + int(parts[2])
            except (ValueError, TypeError):
                pass

        # Cover image
        pic = v.get('pic', '')
        if pic and pic.startswith('//'):
            pic = 'https:' + pic

        # Base stats from user post videos API
        play_count = v.get('play', 0) or 0
        comment_count = v.get('comment', 0) or 0
        danmaku_count = v.get('video_review', 0) or 0
        like_count = 0
        collect_count = 0
        share_count = 0

        # Fetch detailed stats (like, favorite, share) from video detail API
        # Limit to first 15 videos to avoid too many API calls
        if aid and len(videos) < 15:
            try:
                detail = tikhub_get('/api/v1/bilibili/web/fetch_video_detail', TIKHUB_API_KEY,
                                    params={'aid': str(aid)}, timeout=15)
                stat = detail.get('data', {}).get('data', {}).get('View', {}).get('stat', {})
                if isinstance(stat, dict):
                    like_count = stat.get('like', 0) or 0
                    collect_count = stat.get('favorite', 0) or 0
                    share_count = stat.get('share', 0) or 0
                    # Use detailed stats if more accurate
                    if stat.get('view'):
                        play_count = stat.get('view', 0) or 0
                    if stat.get('reply'):
                        comment_count = stat.get('reply', 0) or 0
            except Exception as e:
                print(f'⚠️ B站获取视频 {bvid} 详情失败: {e}')

        videos.append({
            'id': bvid or str(aid),
            'title': title,
            'thumbnail': pic,
            'viewCount': play_count,
            'likeCount': like_count,
            'commentCount': comment_count,
            'shareCount': share_count,
            'collectCount': collect_count,
            'danmakuCount': danmaku_count,
            'duration': duration_sec,
            'durationFormatted': str(duration_str) if duration_str else '',
            'publishedAt': datetime.fromtimestamp(v.get('created', 0)).strftime('%Y-%m-%d %H:%M') if v.get('created') else '',
            'music_title': '',
            'music_author': '',
            'music_id': '',
            'hashtags': [],
            'is_top': False,
            'description': v.get('description', ''),
            'url': f'https://www.bilibili.com/video/{bvid}',
            'platform': 'bilibili',
        })
    return videos


# ==================== Platform: WeChat Channels ====================
def wechat_search_user(keyword):
    """Search wechat channels user by keyword."""
    result = tikhub_get('/api/v1/wechat_channels/fetch_user_search_v2', TIKHUB_API_KEY,
                        params={'keywords': keyword, 'page': 1}, timeout=60)
    items = result.get('data', {}).get('items', [])
    out = []
    for u in items[:10]:
        if not isinstance(u, dict):
            continue
        jump = u.get('jumpInfo', {})
        finder_name = jump.get('userName', '') if isinstance(jump, dict) else ''
        out.append({
            'id': finder_name,
            'name': u.get('nickname', ''),  # May need to extract from desc
            'avatar': u.get('headUrl', ''),
            'fans': u.get('fansCount', u.get('sub_title', '')),
            'desc': u.get('desc', ''),
            'authInfo': u.get('authInfo', ''),
            'docID': u.get('docID', ''),
            'platform': 'wechat',
            'platform_id': finder_name,
            'platform_id_label': '微信号',
        })
    return out


def wechat_get_profile(finder_username):
    """WeChat Channels doesn't have a dedicated profile endpoint.
    Return basic info from search results."""
    # Search for the user
    result = tikhub_get('/api/v1/wechat_channels/fetch_user_search_v2', TIKHUB_API_KEY,
                        params={'keywords': finder_username, 'page': 1}, timeout=60)
    items = result.get('data', {}).get('items', [])
    for u in items:
        if not isinstance(u, dict):
            continue
        jump = u.get('jumpInfo', {})
        if isinstance(jump, dict) and jump.get('userName') == finder_username:
            return {
                'id': finder_username,
                'name': u.get('nickname', finder_username),
                'avatar': u.get('headUrl', ''),
                'signature': u.get('desc', ''),
                'fans': u.get('fansCount', u.get('sub_title', '')),
                'following': 0,
                'likes': 0,
                'posts': 0,
                'authInfo': u.get('authInfo', ''),
                'platform': 'wechat',
            }
    # Fallback: return minimal info
    return {
        'id': finder_username,
        'name': finder_username,
        'avatar': '',
        'signature': '',
        'fans': 0,
        'following': 0,
        'likes': 0,
        'posts': 0,
        'platform': 'wechat',
    }


def wechat_get_videos(finder_username, count=20):
    """WeChat Channels doesn't have a dedicated user videos endpoint.
    Use search to find videos by this user."""
    result = tikhub_get('/api/v1/wechat_channels/fetch_search_ordinary', TIKHUB_API_KEY,
                        params={'keywords': finder_username}, timeout=60)
    items = result.get('data', {}).get('items', [])
    if not items:
        # Fallback to latest search
        result = tikhub_get('/api/v1/wechat_channels/fetch_search_latest', TIKHUB_API_KEY,
                            params={'keywords': finder_username}, timeout=60)
        items = result.get('data', {}).get('items', [])
    videos = []
    for item in items[:count]:
        if not isinstance(item, dict):
            continue
        source = item.get('source', {})
        # Only include videos from this user
        source_name = source.get('title', '') if isinstance(source, dict) else ''
        # Can't perfectly filter, include all results
        duration_str = item.get('duration', '')
        duration_sec = 0
        if duration_str and ':' in str(duration_str):
            parts = str(duration_str).split(':')
            try:
                if len(parts) == 2:
                    duration_sec = int(parts[0]) * 60 + int(parts[1])
            except (ValueError, TypeError):
                pass

        like_num = item.get('likeNum', '0')
        try:
            like_num = int(like_num)
        except (ValueError, TypeError):
            like_num = 0

        videos.append({
            'id': item.get('hashDocID', item.get('id', '')),
            'title': item.get('title', ''),
            'thumbnail': item.get('image', ''),
            'viewCount': 0,
            'likeCount': like_num,
            'commentCount': 0,
            'shareCount': 0,
            'collectCount': 0,
            'duration': duration_sec,
            'durationFormatted': str(duration_str) if duration_str else '',
            'publishedAt': item.get('dateTime', ''),
            'music_title': '',
            'music_author': '',
            'music_id': '',
            'hashtags': [],
            'is_top': False,
            'channelTitle': source_name,
            'channelAvatar': source.get('iconUrl', '') if isinstance(source, dict) else '',
            'url': item.get('exportId', ''),
            'platform': 'wechat',
        })
    return videos


# ==================== Unified API ====================
PLATFORM_SEARCH = {
    'douyin': douyin_search_user,
    'xiaohongshu': xiaohongshu_search_user,
    'bilibili': bilibili_search_user,
    'wechat': wechat_search_user,
}

PLATFORM_PROFILE = {
    'douyin': douyin_get_profile,
    'xiaohongshu': xiaohongshu_get_profile,
    'bilibili': bilibili_get_profile,
    'wechat': wechat_get_profile,
}

PLATFORM_VIDEOS = {
    'douyin': douyin_get_videos,
    'xiaohongshu': xiaohongshu_get_notes,
    'bilibili': bilibili_get_videos,
    'wechat': wechat_get_videos,
}


# ==================== LLM Analysis ====================
def call_llm(prompt, system_msg=''):
    """Call LLM API for analysis."""
    if not LLM_API_KEY or not LLM_API_URL:
        raise Exception('LLM API 未配置，请在设置中填写 API Key 和 API URL')

    headers = {
        'Authorization': f'Bearer {LLM_API_KEY}',
        'Content-Type': 'application/json',
    }
    messages = []
    if system_msg:
        messages.append({'role': 'system', 'content': system_msg})
    messages.append({'role': 'user', 'content': prompt})

    body = {
        'model': LLM_MODEL,
        'messages': messages,
        'temperature': 0.7,
        'max_tokens': 8000,
    }

    r = req_lib.post(LLM_API_URL, headers=headers, json=body, timeout=120,
                     proxies={'http': PROXY_URL, 'https': PROXY_URL} if USE_PROXY else None)
    if r.status_code != 200:
        raise Exception(f'LLM API 错误: HTTP {r.status_code} - {r.text[:200]}')
    result = r.json()
    content = result.get('choices', [{}])[0].get('message', {}).get('content', '')
    return content


# ==================== Routes ====================
@app.route('/')
def index():
    return render_template('index.html')


@app.route('/api/settings', methods=['GET'])
def get_settings():
    return jsonify({
        'tikhub_key_set': bool(TIKHUB_API_KEY),
        'tikhub_key_masked': TIKHUB_API_KEY[:6] + '***' + TIKHUB_API_KEY[-4:] if len(TIKHUB_API_KEY) > 10 else ('***' if TIKHUB_API_KEY else ''),
        'llm_key_set': bool(LLM_API_KEY),
        'llm_key_masked': LLM_API_KEY[:6] + '***' + LLM_API_KEY[-4:] if len(LLM_API_KEY) > 10 else ('***' if LLM_API_KEY else ''),
        'llm_api_url': LLM_API_URL,
        'llm_model': LLM_MODEL,
    })


@app.route('/api/settings', methods=['POST'])
def save_settings():
    global TIKHUB_API_KEY, LLM_API_KEY, LLM_API_URL, LLM_MODEL
    data = request.json or {}
    if data.get('tikhub_key'):
        TIKHUB_API_KEY = data['tikhub_key']
        os.environ['TIKHUB_API_KEY'] = TIKHUB_API_KEY
    if data.get('llm_key'):
        LLM_API_KEY = data['llm_key']
        os.environ['LLM_API_KEY'] = LLM_API_KEY
    if data.get('llm_api_url'):
        LLM_API_URL = data['llm_api_url']
        os.environ['LLM_API_URL'] = LLM_API_URL
    if data.get('llm_model'):
        LLM_MODEL = data['llm_model']
        os.environ['LLM_MODEL'] = LLM_MODEL
    _save_settings_to_file()
    return jsonify({'success': True})


@app.route('/api/search_user', methods=['POST'])
def search_user():
    """Search for a user on a specific platform."""
    data = request.json or {}
    keyword = data.get('keyword', '').strip()
    platform = data.get('platform', '').strip()

    if not keyword:
        return jsonify({'error': '请输入账号名称'}), 400
    if platform not in PLATFORM_SEARCH:
        return jsonify({'error': f'不支持的平台: {platform}'}), 400
    if not TIKHUB_API_KEY:
        return jsonify({'error': 'TikHub API Key 未配置'}), 400

    try:
        users = PLATFORM_SEARCH[platform](keyword)
        return jsonify({'users': users})
    except TikHubError as e:
        return jsonify({'error': str(e)}), 400
    except Exception as e:
        traceback.print_exc()
        return jsonify({'error': f'搜索失败: {str(e)}'}), 500


@app.route('/api/analyze', methods=['POST'])
def analyze_competitor():
    """Analyze a competitor: get profile + videos, then AI analysis."""
    data = request.json or {}
    platform = data.get('platform', '').strip()
    user_id = data.get('user_id', '').strip()
    user_name = data.get('user_name', '').strip()
    video_count = min(int(data.get('video_count', 20)), 50)

    if not platform or not user_id:
        return jsonify({'error': '请提供平台和用户ID'}), 400
    if platform not in PLATFORM_PROFILE:
        return jsonify({'error': f'不支持的平台: {platform}'}), 400
    if not TIKHUB_API_KEY:
        return jsonify({'error': 'TikHub API Key 未配置'}), 400

    try:
        # 1. Get profile
        profile = PLATFORM_PROFILE[platform](user_id)

        # 2. Get videos
        videos = PLATFORM_VIDEOS[platform](user_id, count=video_count)

        # Fix posts count if API doesn't provide it (e.g. xiaohongshu, bilibili)
        if not profile.get('posts'):
            profile['posts'] = len(videos)

        return jsonify({
            'profile': profile,
            'videos': videos,
        })
    except TikHubError as e:
        return jsonify({'error': str(e)}), 400
    except Exception as e:
        traceback.print_exc()
        return jsonify({'error': f'分析失败: {str(e)}'}), 500


@app.route('/api/ai_analyze', methods=['POST'])
def ai_analyze():
    """AI analysis of competitor data."""
    data = request.json or {}
    profile = data.get('profile', {})
    videos = data.get('videos', [])
    platform = data.get('platform', '')

    if not LLM_API_KEY:
        return jsonify({'error': 'LLM API 未配置，请在设置中填写'}), 400

    try:
        # Sort videos by likes to find top videos
        sorted_videos = sorted(videos, key=lambda v: int(v.get('likeCount', 0) or 0), reverse=True)
        top5 = sorted_videos[:5]

        # Build prompt
        platform_names = {'douyin': '抖音', 'xiaohongshu': '小红书', 'bilibili': 'B站', 'wechat': '视频号'}
        platform_cn = platform_names.get(platform, platform)

        video_summary = ""
        for i, v in enumerate(top5, 1):
            video_summary += f"""
视频{i}: {v.get('title', '无标题')}
- 点赞: {v.get('likeCount', 0)} | 评论: {v.get('commentCount', 0)} | 分享: {v.get('shareCount', 0)} | 收藏: {v.get('collectCount', 0)}
- 时长: {v.get('durationFormatted', 'N/A')} | 发布时间: {v.get('publishedAt', 'N/A')}
- 音乐: {v.get('music_title', 'N/A')} - {v.get('music_author', 'N/A')}
- 话题标签: {', '.join(v.get('hashtags', [])) or '无'}
"""

        prompt = f"""你是一位资深的内容营销分析师，请对以下{platform_cn}竞品账号进行**极其深入**的分析。

## 账号基本信息
- 平台: {platform_cn}
- 昵称: {profile.get('name', 'N/A')}
- 粉丝数: {profile.get('fans', 'N/A')}
- 获赞数: {profile.get('likes', 'N/A')}
- 作品数: {profile.get('posts', 'N/A')}
- 简介: {profile.get('signature', 'N/A')}
- IP归属地: {profile.get('ip_location', 'N/A')}

## 爆款视频 TOP5（按点赞排序）
{video_summary}

## 全部视频数据概览
总视频数: {len(videos)}
平均点赞: {sum(int(v.get('likeCount', 0) or 0) for v in videos) // max(len(videos), 1)}
平均评论: {sum(int(v.get('commentCount', 0) or 0) for v in videos) // max(len(videos), 1)}
平均收藏: {sum(int(v.get('collectCount', 0) or 0) for v in videos) // max(len(videos), 1)}
互动率(评论+分享/点赞): {round(sum(int(v.get('commentCount', 0) or 0) + int(v.get('shareCount', 0) or 0) for v in videos) / max(sum(int(v.get('likeCount', 0) or 0) for v in videos), 1) * 100, 1)}%

请输出以下分析报告（用JSON格式）：

```json
{{
  "content_strategy": {{
    "summary": "内容策略总结（100字以内，需包含定位、风格、核心差异化）",
    "content_types": ["内容类型1（占比估算）", "内容类型2（占比估算）", "内容类型3（占比估算）"],
    "posting_frequency": "发布频率分析（具体到每周几篇、什么时间段发布）",
    "content_pillars": ["内容支柱1：说明", "内容支柱2：说明", "内容支柱3：说明"],
    "target_audience": "目标受众画像（年龄/性别/兴趣/痛点）",
    "tone_and_style": "语言风格与调性分析（口语化程度、情感色彩、专业度）"
  }},
  "top_videos_analysis": [
    {{
      "title": "视频标题",
      "why_viral": "爆款原因深度分析（80字以上，需从选题角度、情绪触发、社交货币、算法友好度等维度分析）",
      "copy_structure": "文案结构详细拆解：\\n1️⃣ 前3秒Hook：具体写什么内容、用了什么技巧（悬念/冲突/反常识/痛点戳中）\\n2️⃣ 内容展开（5-15秒）：叙事方式、节奏控制、信息密度\\n3️⃣ 高潮/转折点：核心信息如何呈现\\n4️⃣ 结尾CTA：引导什么行为、用什么话术",
      "rewrite_copy": "仿写文案（保持结构但换主题，200字以上，需完整包含Hook→展开→高潮→CTA四段，语气自然有代入感，符合{platform_cn}平台调性）",
      "key_techniques": ["技巧1：具体说明", "技巧2：具体说明", "技巧3：具体说明"]
    }}
  ],
  "audio_analysis": {{
    "summary": "音效/音乐使用策略深度分析（BGM选择逻辑、音效与内容配合方式）",
    "recommendations": ["音乐推荐1：具体说明为什么推荐", "音乐推荐2：具体说明为什么推荐", "音乐推荐3：具体说明为什么推荐"],
    "audio_patterns": "该账号音频使用的规律总结"
  }},
  "data_insights": {{
    "engagement_pattern": "互动数据规律分析（哪类内容互动率高、点赞/评论/收藏比例说明什么）",
    "content_gap": "内容空白区（该账号没做但受众可能需要的方向）",
    "growth_potential": "增长潜力评估（基于当前数据的趋势判断）"
  }},
  "strengths": ["优势1：具体说明为什么是优势", "优势2：具体说明为什么是优势", "优势3：具体说明为什么是优势", "优势4"],
  "weaknesses": ["劣势1：具体说明为什么是劣势及改进方向", "劣势2：具体说明为什么是劣势及改进方向", "劣势3"],
  "opportunities": ["机会1：具体说明为什么是机会及如何抓住", "机会2：具体说明为什么是机会及如何抓住", "机会3：具体说明为什么是机会及如何抓住"],
  "action_items": ["行动建议1：具体可执行的步骤", "行动建议2：具体可执行的步骤", "行动建议3：具体可执行的步骤", "行动建议4：具体可执行的步骤", "行动建议5：具体可执行的步骤", "行动建议6"]
}}
```

请确保：
1. 分析要有深度和独到见解，不要泛泛而谈，每个观点都要有数据支撑
2. 爆款原因分析要从选题、情绪、社交传播、算法推荐等多维度深入拆解
3. 文案结构拆解要非常具体，精确到每个段落的写作技巧和目的
4. 仿写文案要完整、自然、有代入感，200字以上，不能只是大纲或要点
5. 仿写文案要符合{platform_cn}平台风格和用户阅读习惯
6. 数据洞察部分要基于提供的具体数字进行分析，不做无依据推测
7. 只输出JSON，不要有其他文字"""

        system_msg = f"你是一位专业的{platform_cn}内容营销分析师，擅长竞品分析和内容策略制定。请用中文回答。"

        result = call_llm(prompt, system_msg)

        # Parse JSON from LLM response
        json_match = re.search(r'```json\s*(.*?)\s*```', result, re.DOTALL)
        if json_match:
            analysis = json.loads(json_match.group(1))
        else:
            # Try to parse the whole response as JSON
            try:
                analysis = json.loads(result)
            except json.JSONDecodeError:
                analysis = {'raw_analysis': result}

        # Merge real video data (thumbnail, stats) into top_videos_analysis
        top_analysis = analysis.get('top_videos_analysis', [])
        for i, va in enumerate(top_analysis):
            if i < len(top5):
                v = top5[i]
                va['thumbnail'] = v.get('thumbnail', '')
                va['likeCount'] = v.get('likeCount', 0)
                va['commentCount'] = v.get('commentCount', 0)
                va['shareCount'] = v.get('shareCount', 0)
                va['collectCount'] = v.get('collectCount', 0)
                va['viewCount'] = v.get('viewCount', 0)
                va['durationFormatted'] = v.get('durationFormatted', '')
                va['publishedAt'] = v.get('publishedAt', '')

        return jsonify({'analysis': analysis})

    except json.JSONDecodeError as e:
        return jsonify({'error': f'AI返回数据解析失败: {str(e)}', 'raw': result[:500]}), 400
    except Exception as e:
        traceback.print_exc()
        return jsonify({'error': f'AI分析失败: {str(e)}'}), 500


@app.route('/api/batch_analyze', methods=['POST'])
def batch_analyze():
    """Batch analyze multiple competitors."""
    data = request.json or {}
    tasks = data.get('tasks', [])

    if not tasks:
        return jsonify({'error': '请提供分析任务'}), 400
    if not TIKHUB_API_KEY:
        return jsonify({'error': 'TikHub API Key 未配置'}), 400

    results = []
    for task in tasks:
        platform = task.get('platform', '').strip()
        user_id = task.get('user_id', '').strip()
        user_name = task.get('user_name', '').strip()

        if not platform or not user_id:
            results.append({'error': f'缺少平台或用户ID: {user_name}', 'name': user_name})
            continue

        try:
            profile = PLATFORM_PROFILE[platform](user_id)
            videos = PLATFORM_VIDEOS[platform](user_id, count=20)
            results.append({
                'profile': profile,
                'videos': videos,
                'name': user_name or profile.get('name', ''),
            })
        except Exception as e:
            results.append({'error': str(e), 'name': user_name})

    return jsonify({'results': results})


@app.route('/api/export_report', methods=['POST'])
def export_report():
    """Export analysis report to Excel."""
    data = request.json or {}
    profile = data.get('profile', {})
    videos = data.get('videos', [])
    analysis = data.get('analysis', {})
    platform = data.get('platform', '')

    platform_names = {'douyin': '抖音', 'xiaohongshu': '小红书', 'bilibili': 'B站', 'wechat': '视频号'}

    wb = Workbook()

    # Sheet 1: Profile
    ws1 = wb.active
    ws1.title = '账号概况'
    thin = Side(border_style='thin')
    header_font = Font(bold=True, size=12)
    headers = ['指标', '数据']
    for i, h in enumerate(headers, 1):
        c = ws1.cell(row=1, column=i, value=h)
        c.font = header_font
        c.fill = PatternFill(start_color='E8F0FE', fill_type='solid')

    rows = [
        ('平台', platform_names.get(platform, platform)),
        ('昵称', profile.get('name', '')),
        ('粉丝数', str(profile.get('fans', ''))),
        ('获赞数', str(profile.get('likes', ''))),
        ('作品数', str(profile.get('posts', ''))),
        ('简介', profile.get('signature', '')),
    ]
    for i, (k, v) in enumerate(rows, 2):
        ws1.cell(row=i, column=1, value=k)
        ws1.cell(row=i, column=2, value=str(v))
    ws1.column_dimensions['A'].width = 15
    ws1.column_dimensions['B'].width = 60

    # Sheet 2: Videos
    ws2 = wb.create_sheet('视频数据')
    v_headers = ['标题', '点赞', '评论', '分享', '收藏', '时长', '发布时间', '链接']
    for i, h in enumerate(v_headers, 1):
        c = ws2.cell(row=1, column=i, value=h)
        c.font = header_font
        c.fill = PatternFill(start_color='E8F0FE', fill_type='solid')
    for vi, v in enumerate(videos, 2):
        ws2.cell(row=vi, column=1, value=v.get('title', '')[:100])
        ws2.cell(row=vi, column=2, value=str(v.get('likeCount', 0)))
        ws2.cell(row=vi, column=3, value=str(v.get('commentCount', 0)))
        ws2.cell(row=vi, column=4, value=str(v.get('shareCount', 0)))
        ws2.cell(row=vi, column=5, value=str(v.get('collectCount', 0)))
        ws2.cell(row=vi, column=6, value=v.get('durationFormatted', ''))
        ws2.cell(row=vi, column=7, value=v.get('publishedAt', ''))
        ws2.cell(row=vi, column=8, value=v.get('url', ''))

    # Sheet 3: AI Analysis
    if analysis:
        ws3 = wb.create_sheet('AI分析报告')
        ws3.column_dimensions['A'].width = 20
        ws3.column_dimensions['B'].width = 80

        row = 1
        def write_section(title, content, row):
            ws3.cell(row=row, column=1, value=title).font = Font(bold=True, size=11)
            row += 1
            if isinstance(content, list):
                for item in content:
                    if isinstance(item, dict):
                        for k, v in item.items():
                            ws3.cell(row=row, column=1, value=k)
                            ws3.cell(row=row, column=2, value=str(v)[:200])
                            row += 1
                        row += 1  # blank row between items
                    else:
                        ws3.cell(row=row, column=1, value='')
                        ws3.cell(row=row, column=2, value=str(item)[:200])
                        row += 1
            elif isinstance(content, str):
                ws3.cell(row=row, column=2, value=content)
                row += 1
            row += 1  # blank row between sections
            return row

        cs = analysis.get('content_strategy', {})
        if cs:
            row = write_section('内容策略', [
                f"总结: {cs.get('summary', '')}",
                f"内容类型: {', '.join(cs.get('content_types', []))}",
                f"发布频率: {cs.get('posting_frequency', '')}",
                f"内容支柱: {', '.join(cs.get('content_pillars', []))}",
            ], row)

        top = analysis.get('top_videos_analysis', [])
        if top:
            row = write_section('爆款视频分析', top, row)

        audio = analysis.get('audio_analysis', {})
        if audio:
            row = write_section('音效分析', [
                f"总结: {audio.get('summary', '')}",
                f"推荐: {'; '.join(audio.get('recommendations', []))}",
            ], row)

        for key, label in [('strengths', '优势'), ('weaknesses', '劣势'), ('opportunities', '机会'), ('action_items', '行动建议')]:
            items = analysis.get(key, [])
            if items:
                row = write_section(label, items, row)

    # Save
    tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
    wb.save(tmp.name)
    tmp.close()
    filename = f"竞品分析_{profile.get('name', 'report')}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return send_file(tmp.name, as_attachment=True, download_name=filename)


# ==================== Main ====================
if __name__ == '__main__':
    print(f'🚀 竞品分析工具启动: http://localhost:{APP_PORT}')
    app.run(host='0.0.0.0', port=APP_PORT, debug=False)
